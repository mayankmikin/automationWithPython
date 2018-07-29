#import openpyxl
#wb = openpyxl.load_workbook('example.xlsx')
#wb.get_sheet_names()
import openpyxl
import os

if __name__ == '__main__':
    os.chdir("X:/sravan/automationWithpython_git")
    print (os.getcwd())
    wb = openpyxl.Workbook()
    #ws1 = wb.create_sheet("Mysheet")
    wb = openpyxl.load_workbook('Sampleformat.xlsx')
    print(wb.get_sheet_names())
    sheet = wb.get_sheet_by_name('Sheet1')
    #sheet['A1']
    for rowOfCellObjects in sheet['A1':'E24']:
        #print('---ROW START---')
        colname=[]
        colval=[]
        for cellObj in rowOfCellObjects:
            #print(cellObj.coordinate, cellObj.value)
            colname.append(cellObj.coordinate)
            colval.append(cellObj.value)
        #print(colname)
        print(colval)
