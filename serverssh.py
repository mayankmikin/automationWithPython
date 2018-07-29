#import  jumpssh
from jumpssh import SSHSession
import openpyxl
import os

# gateway_session = SSHSession('javabatchprd1','c646774', password='Raghuma7up@').open()
# remote_session = gateway_session.get_remote_session('javabatchprd1', password='Raghuma7up@')
#
# remote_session.get('/app/logs/member-portal/pinRequestJob/batch.log','C:/Users/c646774/Downloads/automationWithPython-master/')
print('file saved')
def getProcessewdRecords(jobname_):
    log_file = open('X:/sravan/automationWithpython_git/batch.log')

    A=[]
    records_founds=""
    for line in log_file.readlines():
        #print(line)
        if('DEBUG') in line:
            if('Found') in line:
                records_found=line.split('Found')
                records_found=records_found[len(records_found)-1].split('records')
                records_founds=records_found[0]
            if('Marking personId:') in line:
                str_arr=line.split('Marking personId:')
                #print(str_arr)
                str_2=str_arr[len(str_arr)-1].split('as')
                #print(str_2[0])
                A.append(str_2[0])
    print('processed files for ',jobname_,' are ',len(A))
    return len(A)
#print(A)
#print("records found are: ",records_founds)
#print("processed records are: ",len(A))

# before manipulating xls file lets define some variable
total_columns=5
job_names_arr=['PIN Request ','eEOB Notification','eEOB Email Update','eEOB Fallback ',
               'EID','WebMD SecuredMsg']

if __name__ == '__main__':
    os.chdir("X:/sravan/automationWithpython_git")
    print (os.getcwd())
    wb = openpyxl.Workbook()
    #ws1 = wb.create_sheet("Mysheet")
    wb = openpyxl.load_workbook('Sampleformat.xlsx')
    print(wb.get_sheet_names())
    sheet = wb.get_sheet_by_name('Sheet1')
    #sheet['A1']
    sheet['C3']=getProcessewdRecords(job_names_arr[0])
    for rowOfCellObjects in sheet['A1':'E24']:
        #print('---ROW START---')
        colname=[]
        colval=[]
        for cellObj in rowOfCellObjects:
            #print(cellObj.coordinate, cellObj.value)
            colname.append(cellObj.coordinate)
            colval.append(cellObj.value)
        #print(colname)
        print(colval)
    wb.save('Sampleformat.xlsx')

#grep 2018-07-28 batch.log | more
